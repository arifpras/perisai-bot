"""File upload handlers for document and image analysis via /kei and /kin personas.

Supports:
- Images (JPG, PNG) - OCR text extraction
- PDFs - text extraction
- Text files - direct processing
- Excel sheets - data extraction

Usage:
    Add to create_telegram_app() in telegram_bot.py:
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
"""

import io
import logging
import time
from typing import Optional, Tuple
from pathlib import Path

from telegram import Update
from telegram.ext import ContextTypes
from telegram.error import BadRequest

from document_history import get_analysis_db

logger = logging.getLogger(__name__)

# Optional imports - graceful degradation if not installed
try:
    import pytesseract
    from PIL import Image
    HAS_OCR = True
except ImportError:
    HAS_OCR = False
    logger.warning("pytesseract/PIL not installed - image OCR disabled")

try:
    import PyPDF2
    HAS_PDF = True
except ImportError:
    HAS_PDF = False
    logger.warning("PyPDF2 not installed - PDF processing disabled")

try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    logger.warning("openpyxl not installed - Excel processing disabled")


# Constants
MAX_FILE_SIZE_MB = 10  # 10MB limit
MAX_TEXT_LENGTH = 8000  # Max chars to send to LLM


def is_user_authorized(user_id: int) -> bool:
    """Check if user is authorized. Import from telegram_bot if needed."""
    # This should be imported from telegram_bot
    # For now, return True - adjust to match your auth logic
    return True


async def ask_kei(question: str, dual_mode: bool = False) -> str:
    """Ask Kei persona. Should be imported from telegram_bot."""
    pass


async def ask_kin(question: str, dual_mode: bool = False) -> str:
    """Ask Kin persona. Should be imported from telegram_bot."""
    pass


def extract_text_from_image(image_bytes: bytes) -> Optional[str]:
    """Extract text from image using OCR.
    
    Args:
        image_bytes: Raw image data
        
    Returns:
        Extracted text or None if OCR unavailable
    """
    if not HAS_OCR:
        return None
    
    try:
        image = Image.open(io.BytesIO(image_bytes))
        text = pytesseract.image_to_string(image)
        return text.strip() if text else None
    except Exception as e:
        logger.error(f"OCR extraction failed: {e}")
        return None


def extract_text_from_pdf(file_bytes: bytes) -> Optional[str]:
    """Extract text from PDF file.
    
    Args:
        file_bytes: Raw PDF data
        
    Returns:
        Extracted text or None if extraction fails
    """
    if not HAS_PDF:
        return None
    
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
        text_parts = []
        
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            if text:
                text_parts.append(text)
        
        return "\n".join(text_parts).strip() if text_parts else None
    except Exception as e:
        logger.error(f"PDF extraction failed: {e}")
        return None


def extract_text_from_excel(file_bytes: bytes) -> Optional[str]:
    """Extract text/data from Excel file.
    
    Args:
        file_bytes: Raw Excel data
        
    Returns:
        Extracted data as formatted text or None
    """
    if not HAS_EXCEL:
        return None
    
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                text_parts.append(row_text)
        
        return "\n".join(text_parts).strip() if text_parts else None
    except Exception as e:
        logger.error(f"Excel extraction failed: {e}")
        return None


def truncate_text(text: str, max_length: int = MAX_TEXT_LENGTH) -> str:
    """Truncate text to max length with ellipsis."""
    if len(text) > max_length:
        return text[:max_length] + f"\n\n[Text truncated - {len(text) - max_length} chars omitted]"
    return text


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle photo uploads for analysis by Kei or Kin.
    
    User can:
    1. Send photo with caption like: "/kei analyze this screenshot"
    2. Send photo with caption like: "/kin what's in this image"
    """
    user_id = update.message.from_user.id
    username = update.message.from_user.username or f"user_{user_id}"
    start_time = time.time()
    
    if not is_user_authorized(user_id):
        await update.message.reply_text("⛔ Access denied. This bot is restricted to authorized users only.")
        logger.warning(f"Unauthorized photo access attempt from user_id={user_id}")
        return
    
    try:
        # Get the highest resolution photo
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        
        # Download image
        image_bytes = await file.download_as_bytearray()
        
        # Get caption as question
        question = update.message.caption or "Analyze this image"
        
        # Determine persona from caption prefix
        persona = "kei"  # default
        if question.lower().startswith("/kin"):
            persona = "kin"
            question = question[4:].strip()
        elif question.lower().startswith("/kei"):
            persona = "kei"
            question = question[4:].strip()
        
        # Extract text from image
        extracted_text = extract_text_from_image(image_bytes)
        
        if extracted_text:
            # Combine question with extracted text
            full_prompt = f"Image Analysis Request: {question}\n\nExtracted text from image:\n{extracted_text}"
        else:
            full_prompt = f"Image Analysis Request: {question}\n\n[Note: OCR not available or image contains no readable text]"
        
        # Truncate if needed
        full_prompt = truncate_text(full_prompt)
        
        # Show typing indicator
        try:
            await context.bot.send_chat_action(chat_id=update.message.chat_id, action="typing")
        except Exception as e:
            logger.warning(f"Failed to send typing indicator: {e}")
        
        # Ask appropriate persona
        if persona == "kin":
            answer = await ask_kin(full_prompt)
            query_type = "image_kin"
        else:
            answer = await ask_kei(full_prompt)
            query_type = "image_kei"
        
        if answer and answer.strip():
            try:
                await update.message.reply_text(answer, parse_mode="HTML")
            except BadRequest:
                # Fallback to plain text if HTML parsing fails
                await update.message.reply_text(answer, parse_mode="Markdown")
            
            response_time = time.time() - start_time
            
            # Log to database
            try:
                db = get_analysis_db()
                db.log_analysis(
                    user_id=user_id,
                    username=username,
                    document_name="image.jpg",
                    original_question=question[:200],
                    extracted_preview=extracted_text[:500] if extracted_text else "",
                    analysis_result=answer[:2000],
                    persona=persona,
                    document_type="image",
                    processing_time_ms=response_time * 1000,
                    status="success"
                )
            except Exception as db_err:
                logger.warning(f"Failed to log to database: {db_err}")
            
            # Log metric if available
            try:
                from metrics import metrics
                metrics.log_query(user_id, username, question[:100], query_type, response_time, True, persona=persona)
            except Exception:
                pass
        else:
            await update.message.reply_text("⚠️ Empty response. Please try again.")
    
    except Exception as e:
        logger.error(f"Error handling photo: {e}", exc_info=True)
        await update.message.reply_text(f"❌ Error processing image: {str(e)[:100]}")


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document uploads (PDF, DOCX, TXT, XLSX) for analysis.
    
    User can:
    1. Send file with caption like: "/kei summarize this document"
    2. Send file with caption like: "/kin analyze this spreadsheet"
    """
    user_id = update.message.from_user.id
    username = update.message.from_user.username or f"user_{user_id}"
    start_time = time.time()
    
    if not is_user_authorized(user_id):
        await update.message.reply_text("⛔ Access denied. This bot is restricted to authorized users only.")
        logger.warning(f"Unauthorized document access attempt from user_id={user_id}")
        return
    
    try:
        document = update.message.document
        
        # Check file size
        if document.file_size > MAX_FILE_SIZE_MB * 1024 * 1024:
            await update.message.reply_text(f"❌ File too large. Maximum size: {MAX_FILE_SIZE_MB}MB")
            return
        
        # Get file
        file = await context.bot.get_file(document.file_id)
        file_bytes = await file.download_as_bytearray()
        
        # Get caption as question
        question = update.message.caption or f"Analyze this {document.mime_type} document"
        
        # Determine persona from caption prefix
        persona = "kei"  # default
        if question.lower().startswith("/kin"):
            persona = "kin"
            question = question[4:].strip()
        elif question.lower().startswith("/kei"):
            persona = "kei"
            question = question[4:].strip()
        
        # Extract text based on file type
        extracted_text = None
        file_type_msg = ""
        
        if document.mime_type == "text/plain":
            try:
                extracted_text = file_bytes.decode('utf-8')
                file_type_msg = "Text file"
            except Exception as e:
                logger.error(f"Plain text decode failed: {e}")
                await update.message.reply_text("❌ Could not read text file encoding.")
                return
        
        elif document.mime_type == "application/pdf":
            if not HAS_PDF:
                await update.message.reply_text("⚠️ PDF support not installed. Please upload a different format.")
                return
            extracted_text = extract_text_from_pdf(file_bytes)
            file_type_msg = "PDF document"
        
        elif document.mime_type in [
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ]:
            if not HAS_EXCEL:
                await update.message.reply_text("⚠️ Excel support not installed. Please upload a different format.")
                return
            extracted_text = extract_text_from_excel(file_bytes)
            file_type_msg = "Excel spreadsheet"
        
        elif document.mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            await update.message.reply_text("⚠️ DOCX format not yet supported. Please convert to PDF or TXT.")
            return
        
        else:
            await update.message.reply_text(f"❌ Unsupported file type: {document.mime_type}")
            return
        
        if not extracted_text:
            await update.message.reply_text(f"⚠️ Could not extract text from {file_type_msg}.")
            return
        
        # Show typing indicator
        try:
            await context.bot.send_chat_action(chat_id=update.message.chat_id, action="typing")
        except Exception as e:
            logger.warning(f"Failed to send typing indicator: {e}")
        
        # Combine question with extracted text
        full_prompt = f"Document Analysis Request ({file_type_msg}): {question}\n\nDocument content:\n{extracted_text}"
        full_prompt = truncate_text(full_prompt)
        
        # Ask appropriate persona
        if persona == "kin":
            answer = await ask_kin(full_prompt)
            query_type = "doc_kin"
        else:
            answer = await ask_kei(full_prompt)
            query_type = "doc_kei"
        
        if answer and answer.strip():
            try:
                await update.message.reply_text(answer, parse_mode="HTML")
            except BadRequest:
                # Fallback to plain text if HTML parsing fails
                await update.message.reply_text(answer, parse_mode="Markdown")
            
            response_time = time.time() - start_time
            
            # Log to database
            try:
                db = get_analysis_db()
                db.log_analysis(
                    user_id=user_id,
                    username=username,
                    document_name=document.file_name,
                    original_question=question[:200],
                    extracted_preview=extracted_text[:500] if extracted_text else "",
                    analysis_result=answer[:2000],
                    persona=persona,
                    document_type=file_type_msg.lower().split()[0],  # 'pdf', 'excel', 'text'
                    processing_time_ms=response_time * 1000,
                    status="success"
                )
            except Exception as db_err:
                logger.warning(f"Failed to log to database: {db_err}")
            
            # Log metric if available
            try:
                from metrics import metrics
                metrics.log_query(user_id, username, question[:100], query_type, response_time, True, persona=persona)
            except Exception:
                pass
        else:
            await update.message.reply_text("⚠️ Empty response. Please try again.")
    
    except Exception as e:
        logger.error(f"Error handling document: {e}", exc_info=True)
        
        # Log error to database
        try:
            db = get_analysis_db()
            db.log_analysis(
                user_id=user_id,
                username=username,
                document_name=document.file_name if 'document' in locals() else "unknown",
                original_question=question[:200] if 'question' in locals() else "",
                extracted_preview="",
                analysis_result="",
                persona="kei",
                document_type="unknown",
                processing_time_ms=(time.time() - start_time) * 1000,
                status="error",
                error_message=str(e)[:200]
            )
        except Exception:
            pass
        
        await update.message.reply_text(f"❌ Error processing document: {str(e)[:100]}")


# Integration example for telegram_bot.py:
"""
In create_telegram_app() function, add these lines:

    # File upload handlers
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))

Then import at the top of telegram_bot.py:
    from file_handlers import handle_photo, handle_document
"""
